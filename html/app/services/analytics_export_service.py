"""
Analytics Export Service - Генерация и экспорт ежемесячной аналитики
"""
import os
import logging
from datetime import datetime, date, timedelta
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.core.database import SessionLocal
from app.core.google_drive_service import drive_service
from app.models import Student, Payment, Group, Attendance, User, AttendanceStatus

logger = logging.getLogger(__name__)

class AnalyticsExportService:
    
    def __init__(self, db: Session):
        self.db = db
        
    def generate_monthly_report(self, year: int, month: int) -> str:
        """
        Генерирует Excel отчет за указанный месяц и возвращает путь к файлу.
        """
        # Создаем книгу Excel
        wb = openpyxl.Workbook()
        
        # 1. Лист "Обзор" (Summary)
        self._create_summary_sheet(wb, year, month)
        
        # 2. Лист "Ученики" (Students)
        self._create_students_sheet(wb)
        
        # 3. Лист "Платежи" (Payments)
        self._create_payments_sheet(wb, year, month)
        
        # 4. Лист "Посещаемость" (Attendance)
        self._create_attendance_sheet(wb, year, month)
        
        # Сохраняем файл
        filename = f"Analytics_{year}_{month:02d}.xlsx"
        filepath = os.path.join("/tmp", filename)
        wb.save(filepath)
        
        return filepath

    def _create_summary_sheet(self, wb, year, month):
        ws = wb.active
        ws.title = "Обзор"
        
        # Стили
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        
        ws['A1'] = f"Отчет за {month:02d}.{year}"
        ws['A1'].font = header_font
        
        # Статистика
        total_students = self.db.query(Student).filter(Student.status == 'active').count()
        total_frozen = self.db.query(Student).filter(Student.is_frozen == True).count()
        
        # Финансы за месяц
        income = self.db.query(func.sum(Payment.amount)).filter(
            func.extract('month', Payment.payment_date) == month,
            func.extract('year', Payment.payment_date) == year,
            Payment.status == 'completed'
        ).scalar() or 0
        
        ws['A3'] = "Ключевые показатели"
        ws['A3'].font = title_font
        
        metrics = [
            ("Активных учеников", total_students),
            ("Замороженных", total_frozen),
            ("Выручка за месяц (MDL)", income),
            ("Дата формирования", datetime.now().strftime("%d.%m.%Y %H:%M"))
        ]
        
        for idx, (label, value) in enumerate(metrics, start=4):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value

        # Автоширина
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_students_sheet(self, wb):
        ws = wb.create_sheet("Ученики")
        
        headers = ["ID", "ФИО", "Группа", "Статус", "Баланс", "Родитель", "Телефон"]
        ws.append(headers)
        
        # Стили заголовка
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
        students = self.db.query(Student).all()
        
        for s in students:
            parent_name = ""
            parent_phone = s.parent_phone or ""
            
            # Пытаемся найти родителя через guardians
            if s.guardians:
                g = s.guardians[0]
                if g.user:
                    parent_name = g.user.full_name
                    parent_phone = g.user.phone or parent_phone
            
            row = [
                s.id,
                f"{s.last_name} {s.first_name}",
                s.group.name if s.group else "-",
                "Заморожен" if s.is_frozen else s.status,
                s.balance,
                parent_name,
                parent_phone
            ]
            ws.append(row)
            
        # Автоширина
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    def _create_payments_sheet(self, wb, year, month):
        ws = wb.create_sheet("Платежи")
        
        headers = ["ID", "Дата", "Ученик", "Группа", "Сумма", "Период оплаты", "Метод", "Статус"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
        payments = self.db.query(Payment).filter(
            func.extract('month', Payment.payment_date) == month,
            func.extract('year', Payment.payment_date) == year
        ).order_by(Payment.payment_date.desc()).all()
        
        for p in payments:
            student_name = f"{p.student.last_name} {p.student.first_name}" if p.student else "Unknown"
            group_name = p.student.group.name if p.student and p.student.group else "-"
            
            row = [
                p.id,
                p.payment_date.strftime("%d.%m.%Y"),
                student_name,
                group_name,
                p.amount,
                p.payment_period.strftime("%m.%Y") if p.payment_period else "-",
                p.method,
                p.status
            ]
            ws.append(row)

    def _create_attendance_sheet(self, wb, year, month):
        ws = wb.create_sheet("Посещаемость")
        
        headers = ["Дата", "Группа", "Ученик", "Статус"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
        attendances = self.db.query(Attendance).join(Attendance.event).filter(
            func.extract('month', Attendance.date) == month,
            func.extract('year', Attendance.date) == year
        ).order_by(Attendance.date.desc()).all()
        
        for a in attendances:
            student_name = f"{a.student.last_name} {a.student.first_name}" if a.student else "Unknown"
            group_name = a.event.group.name if a.event and a.event.group else "-"
            
            row = [
                a.date.strftime("%d.%m.%Y"),
                group_name,
                student_name,
                a.status.value
            ]
            ws.append(row)


def export_monthly_analytics_task(year: int, month: int):
    """
    Фоновая задача: генерация и выгрузка отчета (Sync wrapper for scheduler)
    """
    db = SessionLocal()
    try:
        service = AnalyticsExportService(db)
        
        logger.info(f"📊 Starting analytics export for {month}.{year}")
        
        # 1. Generate Excel
        file_path = service.generate_monthly_report(year, month)
        file_name = f"Analytics_{year}_{month:02d}.xlsx"
        
        # 2. Upload to Drive
        if drive_service.enabled:
            link = drive_service.upload_file(file_path, file_name)
            if link:
                logger.info(f"✅ Report uploaded: {link}")
            else:
                logger.error("❌ Failed to upload report to Drive")
        else:
            logger.warning("⚠️ Drive service disabled, skipping upload")
            
        # 3. Cleanup
        if os.path.exists(file_path):
            os.remove(file_path)
            
    except Exception as e:
        logger.error(f"❌ Analytics export failed: {e}")
    finally:
        db.close()
